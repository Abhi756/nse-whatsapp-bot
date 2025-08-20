# -*- coding: utf-8 -*-
import requests
import pickle
import pandas as pd
import time
import random
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

import os

PICKLE_FILE="prev_data.pkl"

def fetch_option_chain(retries=5, backoff=2):
    base_url = "https://www.nseindia.com"
    api_url = "https://www.nseindia.com/api/option-chain-indices?symbol=BANKNIFTY"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": base_url
    }

    session = requests.Session()
    session.get(base_url, headers=headers, timeout=5)
    time.sleep(1)

    had_failure = False   # <-- track if any failure happened

    for attempt in range(retries):
        try:
            resp = session.get(api_url, headers=headers, timeout=10)
            if resp.status_code == 200 and resp.text.strip():
                if 'application/json' in resp.headers.get("Content-Type", ""):
                    j = resp.json()
                    data = j["records"]["data"]
                    expiry = j["records"]["expiryDates"][0]
                    ce = [e["CE"] for e in data if "CE" in e and e["expiryDate"] == expiry]
                    pe = [e["PE"] for e in data if "PE" in e and e["expiryDate"] == expiry]
                    return pd.DataFrame(ce), pd.DataFrame(pe), had_failure
            else:
                print(f"Attempt {attempt+1} failed, retrying...")
                had_failure = True
        except requests.RequestException as e:
            print(f"Request error: {e}")
            had_failure = True
        time.sleep(backoff * (2 ** attempt))
    return None, None, had_failure


def add_changes(df_new, df_old):
    """Adds Buy Change and Sell Change columns by comparing with previous dataframe."""
    if df_old is None:
        df_new["Buy Change"] = 0
        df_new["Sell Change"] = 0
        return df_new
    df_new = df_new.merge(
        df_old[["strikePrice", "totalBuyQuantity", "totalSellQuantity"]],
        on="strikePrice", how="left", suffixes=("", "_prev")
    )
    df_new["Buy Change"] = df_new["totalBuyQuantity"] - df_new["totalBuyQuantity_prev"].fillna(0)
    df_new["Sell Change"] = df_new["totalSellQuantity"] - df_new["totalSellQuantity_prev"].fillna(0)
    df_new.drop(columns=["totalBuyQuantity_prev", "totalSellQuantity_prev"], inplace=True)
    return df_new
def save_to_excel(file_path, df_ce, df_pe,ignore_prev=False):
    try:
        # Load previous data if exists
        if not ignore_prev and os.path.exists(PICKLE_FILE):
            with open(PICKLE_FILE, "rb") as f:
                prev_data = pickle.load(f)
            prev_ce = prev_data.get("CE")
            prev_pe = prev_data.get("PE")
        else:
            prev_ce = None
            prev_pe = None

        # Add change columns
        df_ce = add_changes(df_ce, prev_ce)
        df_pe = add_changes(df_pe, prev_pe)

        # Save current data for next iteration
        with open(PICKLE_FILE, "wb") as f:
            pickle.dump({"CE": df_ce, "PE": df_pe}, f)

        now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        ce_sheet = f"CE_{now}"
        pe_sheet = f"PE_{now}"

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            wb = load_workbook(file_path) if os.path.exists(file_path) else None

        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a" if os.path.exists(file_path) else "w") as writer:
            df_ce.to_excel(writer, sheet_name=ce_sheet, index=False)
            df_pe.to_excel(writer, sheet_name=pe_sheet, index=False)

        # Apply conditional formatting
        wb = load_workbook(file_path)

        for sheet_name in [ce_sheet, pe_sheet]:
            ws = wb[sheet_name]
            last_col = ws.max_column
            buy_change_col = None
            sell_change_col = None

            # Find columns for Buy Change and Sell Change
            for col in range(1, last_col+1):
                header = ws.cell(row=1, column=col).value
                if header == "Buy Change":
                    buy_change_col = col
                elif header == "Sell Change":
                    sell_change_col = col

            if buy_change_col:
                # Highlight Buy Change > 1000 (green)
                ws.conditional_formatting.add(
                    f"{chr(64+buy_change_col)}2:{chr(64+buy_change_col)}{ws.max_row}",
                    CellIsRule(operator="greaterThan", formula=["1000"], fill=PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"))
                )

            if sell_change_col:
                # Highlight Sell Change > 1000 (red)
                ws.conditional_formatting.add(
                    f"{chr(64+sell_change_col)}2:{chr(64+sell_change_col)}{ws.max_row}",
                    CellIsRule(operator="greaterThan", formula=["1000"], fill=PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"))
                )

        wb.save(file_path)
        print(f"Saved & highlighted: {ce_sheet}, {pe_sheet}")
        print("\n--- CE Sheet Columns ---")
        print(df_ce.columns.tolist())

    except PermissionError as e:
        print(f"Permission error: {e}. Please close the file.")
def run_scheduler():
    start = datetime.strptime("09:15:00", "%H:%M:%S").time()
    end = datetime.strptime("15:30:00", "%H:%M:%S").time()
    path = "optiondata8.xlsx"
    while True:
        now = datetime.now().time()
        if start <= now <= end:
            print(f"Fetching data at {datetime.now()}…")
            df_ce, df_pe,had_failure = fetch_option_chain()
            if df_ce is not None and df_pe is not None:
                save_to_excel(path, df_ce, df_pe,ignore_prev=had_failure)
            time.sleep(55 + random.randint(0,10))
        elif now > end:
            print("Market closed. Scheduler halted.")
            break
        else:
            time.sleep(10)

if __name__ == "__main__":
    run_scheduler()
